from openpyxl import Workbook
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from time import sleep
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# load page for company name
browser.get('https://leetcode.com/company/[COMPANY NAME]]/')
sleep(4)

username = browser.find_element('id', 'id_login')
#add in username
username.send_keys('')

password = browser.find_element('id', 'id_password')
password.send_keys('')
form = browser.find_element('id','signin_btn')
sleep(6)
form.click()

WebDriverWait(browser, 11).until(EC.title_contains("[Company name] - LeetCode"))

sleep(13)

frequency = browser.find_element(By.ID,'app').find_elements(By.TAG_NAME,'th')[5]
sleep(3)
frequency.click()
frequency.click()

list_of_links = []
all_question_content = []

table = browser.find_element(By.XPATH,"//table[@class='table table__XKyc']")
rows = table.find_elements(By.TAG_NAME,'tr')
number_of_rows = len(rows)
for i in range(1,number_of_rows):
    list_of_links.append(rows[i].find_element(By.TAG_NAME,'a').get_attribute('href'))

print(list_of_links)
#browser.find_elements(By.XPATH, "//div[contains(@class, 'content__u3I1 question-content__JfgR')]")[0].text
# questions = browser.find_elements(By.XPATH, "//div[contains(@class, 'content__u3I1 question-content__JfgR')]")

for i in range(len(list_of_links)):
    browser.get(list_of_links[i])
    sleep(4)
    question = browser.find_elements(By.XPATH, "//div[contains(@class, 'content__u3I1 question-content__JfgR')]")[0].text
    all_question_content.append(question)
    # all_question_content.append(questions[0].text)

excelFileName = "[company_name]_leetcode.xlsx"
sheetName = "[company_name] Interview Questions"

df = pd.DataFrame({
    'Question URL':list_of_links,
    'Question Contents': all_question_content
})

wb = Workbook()

sheet1 = wb.create_sheet(sheetName)
sheet1.cell(1,1,'Question Link')
sheet1.cell(1,2,'Question Content')

for i in range(0, df.__len__()):
    sheet1.cell(i+2, 1, df['Question URL'][i])
    sheet1.cell(i + 2, 2, df['Question Contents'][i])

wb.save(excelFileName)
wb.close()
print("Finished excel")